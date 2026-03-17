import os
import openpyxl
import pyperclip
import pyautogui
from time import sleep

pyautogui.FAILSAFE = False

#  NOVA FUNÇÃO 
def clicar_e_preencher(imagem, valor):
    campo = pyautogui.locateCenterOnScreen(imagem, confidence=0.8)
    
    if campo:
        pyautogui.click(campo)
        pyperclip.copy(str(valor))
        pyautogui.hotkey('ctrl', 'v')
    else:
        print(f"❌ {imagem} não encontrado!")

print("🚀 Iniciando automação de cadastro de produtos...")
sleep(1)

# Abrir Chrome
pyautogui.press('win')
pyautogui.write('chrome', interval=0.1)
pyautogui.press('enter')
sleep(2)

# Abrir site
pyautogui.write('https://cadastro-produtos-devaprender.netlify.app/')
pyautogui.press('enter')
sleep(6)

# Focar Chrome
try:
    chrome_windows = pyautogui.getWindowsWithTitle("Chrome")
    if chrome_windows:
        chrome_windows[0].activate()
        print("🌐 Chrome em foco!")
    else:
        pyautogui.hotkey('alt', 'tab')
except:
    pyautogui.hotkey('alt', 'tab')

sleep(2)
pyautogui.alert("Clique em OK quando o site estiver pronto.")

# Caminho da planilha
base_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(base_dir, "produtos_ficticios.xlsx")

print("📂 Caminho da planilha:", file_path)

# Carregar planilha
try:
    workbook = openpyxl.load_workbook(file_path)
    sheet_produtos = workbook["Produtos"]
    print("✅ Planilha carregada!")
except Exception as e:
    print(f"❌ Erro: {e}")
    exit()

# Total de produtos
total = sheet_produtos.max_row - 1

# Loop principal
for i, linha in enumerate(sheet_produtos.iter_rows(min_row=2, values_only=True), start=1):

    nome_produto, descricao_produto, categoria, codigo_ncm, peso, dimensoes, preco, estoque, validade, cor, tamanho, material, fabricante, pais_origem, observacoes, codigo_barras, local_estoque = linha

    print(f"\n📦 [{i}/{total}] Cadastrando: {nome_produto}")

    #  NOVO 
    clicar_e_preencher('campo_nome.png', nome_produto)

    # Descrição (ainda com coordenada)
    pyperclip.copy(descricao_produto)
    pyautogui.click(147,267)
    pyautogui.hotkey('ctrl', 'v')

    # Categoria
    pyperclip.copy(categoria)
    pyautogui.click(143,393)
    pyautogui.hotkey('ctrl', 'v')

    # Código NCM
    pyperclip.copy(codigo_ncm)
    pyautogui.click(142,479)
    pyautogui.hotkey('ctrl', 'v')

    # Peso
    pyperclip.copy(peso)
    pyautogui.click(148,568)
    pyautogui.hotkey('ctrl', 'v')

    # Dimensões
    pyperclip.copy(dimensoes)
    pyautogui.click(146,652)
    pyautogui.hotkey('ctrl', 'v')

    # Próximo
    pyautogui.click(147,700)
    sleep(4)

    # Preço
    pyperclip.copy(preco)
    pyautogui.click(136,199)
    pyautogui.hotkey('ctrl', 'v')

    # Estoque
    pyperclip.copy(estoque)
    pyautogui.click(129,286)
    pyautogui.hotkey('ctrl', 'v')

    # Validade
    pyperclip.copy(validade)
    pyautogui.click(132,373)
    pyautogui.hotkey('ctrl', 'v')

    # Cor
    pyperclip.copy(cor)
    pyautogui.click(134,458)
    pyautogui.hotkey('ctrl', 'v')

    # Tamanho
    pyautogui.click(198,542)
    if tamanho == 'Pequeno':
        pyautogui.click(178,577)
    elif tamanho == 'Medio':
        pyautogui.click(147,606)
    else:
        pyautogui.click(144,638)

    # Material
    pyperclip.copy(material)
    pyautogui.click(136,628)
    pyautogui.hotkey('ctrl', 'v')

    # Próxima página
    pyautogui.click(152,682)
    sleep(4)

    # Fabricante
    pyperclip.copy(fabricante)
    pyautogui.click(231,218)
    pyautogui.hotkey('ctrl', 'v')

    # País
    pyperclip.copy(pais_origem)
    pyautogui.click(132,305)
    pyautogui.hotkey('ctrl', 'v')

    # Observações
    pyperclip.copy(observacoes)
    pyautogui.click(138,390)
    pyautogui.hotkey('ctrl', 'v')

    # Código de barras
    pyperclip.copy(codigo_barras)
    pyautogui.click(131,526)
    pyautogui.hotkey('ctrl', 'v')

    # Local estoque
    pyperclip.copy(local_estoque)
    pyautogui.click(125,613)
    pyautogui.hotkey('ctrl', 'v')

    # Concluir
    pyautogui.click(142,669)
    sleep(4)

    # OK
    pyautogui.click(850,185)
    sleep(3)

    # Finalizar
    pyautogui.click(709,438)

    print(f"✅ Produto {i} cadastrado!")

print("\n🎉 Processo finalizado com sucesso!")
